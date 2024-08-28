from openpyxl import Workbook, load_workbook
from tkinter import *
import cv2
from os.path import join, expanduser
from os import makedirs
from PIL import Image, ImageTk
from tkinter import filedialog
from utilities.models import Models
from processing.process_image import process_image
from utilities.default_conversion_scale import DEFAULT_CONVERSION_SCALE
from pypylon import pylon
from datetime import datetime
import os

width, height = 1300, 700

app = Tk()
app.title("Capillary bridges image processing")

app.bind('<Escape>', lambda e: app.quit())

label_widget = Label(app, width=width, height=height)
label_widget.pack()

empty_image = ImageTk.PhotoImage(Image.new("RGB", (width, height), "white"))
label_widget.configure(image=empty_image)
label_widget.photo_image = empty_image

values_label = Label(app, text="", font=("Helvetica", 10), wraplength=1200, anchor='w')
values_label.pack(pady=10)

running_camera = False
is_recording = False
is_grabbing = False
should_process_image = True

STANDARD_CAMERA = "Standard"
BASLER_CAMERA = "Basler"

BRIGHTEN = "Brighten"
STANDARD_BRIGHTNESS = "Standard brightness"

RECORDINGS_FOLDER = "recordings"
EXCEL_VALUES_FILE_NAME = "measured_values.xlsx"

MIN_EXPOSURE_TIME = 59
MAX_EXPOSURE_TIME = 1000000

base_dir = join(expanduser("~"), "Desktop", "Capillary bridges image analysis")
current_folder_path = None
workbook = None
worksheet = None
image_count = 0
values = None

def create_images_folder_structure():
    global current_folder_path, workbook, worksheet, is_recording

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    current_folder_path = join(base_dir, timestamp)

    if is_recording: current_folder_path = join(base_dir, RECORDINGS_FOLDER, timestamp)

    makedirs(current_folder_path)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Image", "Neck", "Down", "Up", "Left", "Right", "Left major", "Left minor", "Right major", "Right minor", "Left average", "Right average", "Base", "Height", "x", "1/x", "y"])

    workbook.save(join(current_folder_path, EXCEL_VALUES_FILE_NAME))

def show_cam_frame(frame):
    if frame is None: return
    
    opencv_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    
    img = Image.fromarray(opencv_image)
    
    wpercent = (width/float(img.size[0]))
    hsize = int((float(img.size[1])*float(wpercent)))
    resized_image = img.resize((width,hsize), Image.Resampling.LANCZOS)
    
    img = ImageTk.PhotoImage(image=resized_image)
    label_widget.photo_image = img
    label_widget.configure(image=img)

def convert_avi_to_mp4(avi_file_path, output_name):
    os.popen(f"ffmpeg -i {avi_file_path} -c:v mpeg4 {output_name}")
    return True

def open_video(MODEL, selected_camera_index: str, selected_brightness_index: str):
    global running_camera, frame, processed_frame, values

    file_path = filedialog.askopenfilename(title="Select a video file",
                                           filetypes=[("Video files", "*.avi;" "*.mp4;*.mkv;")])

    if file_path:
        video_button_basic.config(state="disabled")

        video_path = file_path
    
        if file_path.split("/")[-1].endswith(".avi"):
            convert_avi_to_mp4(file_path, "temp_video.mp4")

            video_path = "temp_video.mp4"

        cap = cv2.VideoCapture(video_path)

        while not cap.isOpened():
            cap = cv2.VideoCapture(video_path)
            cv2.waitKey(1000)
            print("Wait for the header")

        pos_frame = cap.get(cv2.CAP_PROP_POS_FRAMES)

        scale_entry_value = float(scale_entry.get())

        bright_mode = selected_brightness_index == BRIGHTEN

        while True:
            flag, frame = cap.read()
            if flag:
                processed_frame, _, values = process_image(frame, 0, scale_entry_value, model=Models.NAIVE, bright=bright_mode)

                save_current_frame()

                pos_frame = cap.get(cv2.CAP_PROP_POS_FRAMES)
            else:
                # The next frame is not ready, so we try to read it again
                cap.set(cv2.CAP_PROP_POS_FRAMES, pos_frame-1)
                # It is better to wait for a while for the next frame to be ready
                cv2.waitKey(1000)

            if cap.get(cv2.CAP_PROP_POS_FRAMES) == cap.get(cv2.CAP_PROP_FRAME_COUNT):
                # If the number of captured frames is equal to the total number of frames,
                # we stop
                video_button_basic.config(state="normal")
                os.remove("temp_video.mp4")
                break

def open_image(MODEL, selected_camera_index: str, selected_brightness_index: str):
    global running_camera, frame, processed_frame, values
    if running_camera:
        running_camera = False
        
        if selected_camera_index == STANDARD_CAMERA and basler_camera is not None:
            standard_camera.release()
        elif selected_camera_index == BASLER_CAMERA and basler_camera is not None:
            basler_camera.StopGrabbing()
        
        reset_label()

    file_paths = filedialog.askopenfilenames(title="Select image files",
                                           filetypes=[("Image files", "*.png;*.jpg;*.jpeg;")])
    
    image_button_basic.config(state="disabled")

    for file_path in file_paths:
        frame = cv2.imread(file_path)

        scale_entry_value = float(scale_entry.get())
        
        processed_frame, _, values = process_image(frame, 0, scale_entry_value, './temp', model=MODEL, bright=selected_brightness_index == BRIGHTEN)
        
        save_current_frame()

        # update_values_label(values)

        # if processed_image is not None:
        #     frame = image
        #     processed_frame = processed_image
        #     show_cam_frame(processed_image)
        #     save_button.config(state="normal")

    image_button_basic.config(state="normal")

def open_camera(selected_camera_index: str, selected_brightness_index: str):
    global running_camera, standard_camera, basler_camera, is_grabbing, converter, image_count
    is_grabbing = False
    image_count = 0
    
    #if not is_recording: create_images_folder_structure()
    
    if not running_camera:
        running_camera = True
        
        if selected_camera_index == STANDARD_CAMERA:
            standard_camera = cv2.VideoCapture(0, cv2.CAP_DSHOW)
        elif selected_camera_index == BASLER_CAMERA:
            try:
                basler_camera = pylon.InstantCamera(pylon.TlFactory.GetInstance().CreateFirstDevice())
                exposure_entry.config(state="normal")
                exposure_entry_value = exposure_entry.get()
                if not exposure_entry_value: exposure_entry.insert(0, 5000)
                fps_entry.config(state="normal")
                fps_entry_value = fps_entry.get()
                if not fps_entry_value: fps_entry.insert(0, 60)
            except:
                update_error_message("Basler camera is not found. Try reconnecting it or switch the camera.")
                running_camera = False
                return
            
        capture_camera(selected_camera_index, selected_brightness_index)
        if not is_recording: save_button.config(state="normal")
        
def save_frame(frame, processed_frame, values):
    global workbook, worksheet, image_count, current_folder_path
    
    image_count += 1
    raw_image_filename = f"raw_{image_count}.png"
    processed_image_filename = f"processed_{image_count}.png"
    
    if current_folder_path == None:
        create_images_folder_structure()
        # list_of_files = glob(join(base_dir, '*'))
        # latest_file = max(list_of_files, key=getctime)
        # current_folder_path = latest_file
        
    if worksheet == None:
        workbook = load_workbook(join(current_folder_path, EXCEL_VALUES_FILE_NAME))
        worksheet = workbook.active

    cv2.imwrite(join(current_folder_path, raw_image_filename), frame)
    cv2.imwrite(join(current_folder_path, processed_image_filename), processed_frame)
    
    keys = filter(lambda x: x in values, ["neck", "down", "up", "left", "right", "left major", "left minor", "right major", "right minor", "left average", "right average", "base", "height", "x", "1/x", "y"]) 

    values_list = [values[key] for key in keys]

    worksheet.append([worksheet.max_row, *values_list])
    workbook.save(join(current_folder_path, EXCEL_VALUES_FILE_NAME))
    
def save_current_frame():
    global frame, processed_frame, values, running_camera

    if frame is not None and processed_frame is not None and values is not None:
        save_frame(frame, processed_frame, values)
        if not running_camera: save_button.config(state="disabled")

def capture_standard(selected_brightness_index: str):
    global frame, processed_frame, values, should_process_image
    _, frame = standard_camera.read()
        
    if frame is not None and is_recording:
        #out.write(frame)
        show_cam_frame(frame)

    if frame is not None and not is_recording:
        try:
            scale_entry_value = float(scale_entry.get())

            processed_frame, _, values = process_image(frame, 0, scale_entry_value, model=Models.NAIVE, bright=selected_brightness_index == BRIGHTEN)
            
            update_values_label(values)

            show_cam_frame(processed_frame if should_process_image else frame)
        except:
            print('ERROR: Could not process image')

    label_widget.after(10, lambda: capture_camera(selected_camera_index, selected_brightness_index))
            
def capture_basler(selected_brightness_index: str):
        global is_grabbing, converter, frame, processed_frame, values, should_process_image, basler_writer, c
        if not is_grabbing:
            basler_camera.StartGrabbing(pylon.GrabStrategy_LatestImageOnly)
            is_grabbing = True
            converter = pylon.ImageFormatConverter()
            converter.OutputPixelFormat = pylon.PixelType_BGR8packed
            converter.OutputBitAlignment = pylon.OutputBitAlignment_MsbAligned

        fps_value = None
        fps_entry_value = fps_entry.get()
        if fps_entry_value != '': fps_value = round(float(fps_entry_value), 2)   
        if fps_value:
            basler_camera.AcquisitionFrameRateEnable.Value = True
            basler_camera.AcquisitionFrameRate.Value = fps_value
         
        exposure_time = None
        exposure_entry_value = exposure_entry.get()
        if exposure_entry_value != '': exposure_time = int(exposure_entry_value)   
        if exposure_time and exposure_time >= MIN_EXPOSURE_TIME and exposure_time <= MAX_EXPOSURE_TIME: basler_camera.ExposureTime.SetValue(exposure_time)

        # if not fps_entry_value:
        fps = basler_camera.ResultingFrameRate.Value
        update_frame_rate(fps)
        
        grabResult = basler_camera.RetrieveResult(5000, pylon.TimeoutHandling_ThrowException)
        if grabResult.GrabSucceeded():
            image = converter.Convert(grabResult)
            frame = image.GetArray()

            if frame is not None and is_recording:
                basler_writer.append_data(frame)
                show_cam_frame(frame)
                
            if frame is not None and not is_recording:
                try:
                    scale_entry_value = float(scale_entry.get())

                    processed_frame, _, values = process_image(frame, 0, scale_entry_value, model=Models.NAIVE, bright=selected_brightness_index == BRIGHTEN)

                    update_values_label(values)

                    show_cam_frame(processed_frame if should_process_image else frame)
                except:
                    print('ERROR: Could not process image')

            label_widget.after(10, lambda: capture_camera(selected_camera_index, selected_brightness_index))

        grabResult.Release()

def capture_camera(selected_camera_index: str, selected_brightness_index: str):
    global is_grabbing, basler_writer
    if running_camera:
        if selected_camera_index == STANDARD_CAMERA:
            capture_standard(selected_brightness_index)
        elif selected_camera_index == BASLER_CAMERA:
            capture_basler(selected_brightness_index)
    else:
        if selected_camera_index == STANDARD_CAMERA and standard_camera is not None:
            standard_camera.release()
        elif selected_camera_index == BASLER_CAMERA and basler_camera is not None:
            basler_camera.StopGrabbing()
            is_grabbing = False
        reset_label()
        
def toggle_processing():
    global should_process_image
    should_process_image = not should_process_image

def reset_label():
    label_widget.configure(image=empty_image)
    label_widget.photo_image = empty_image
    save_button.config(state="disabled")
    exposure_entry.config(state="disabled")
    
def update_frame_rate(fps):
    if fps is None: return
    
    formatted_fps = f"FPS: {round(fps, 2)}"
    fps_label.config(text=formatted_fps)
    
def update_error_message(error):
    values_label.config(text=error, font=("Helvetica", 12))

def update_values_label(values):
    if values is None: return

    formatted_values = {key.capitalize(): f'{value:.2f} um' for key, value in values.items()}
    values_text = ", ".join([f"{key}: {value}" for key, value in formatted_values.items()])
    values_label.config(text=values_text, font=("Helvetica", 10))
    
def select_camera(camera_index):
    global selected_camera_index
    selected_camera_index = camera_index

camera_options = [
    STANDARD_CAMERA,
    BASLER_CAMERA
]
selected_camera_index = BASLER_CAMERA 

def select_brightness(brightness_index):
    global selected_brightness_index
    selected_brightness_index = brightness_index

brighten_options = [
    BRIGHTEN,
    STANDARD_BRIGHTNESS
]
selected_brightness_index = STANDARD_BRIGHTNESS 

realtime_button = Button(app, text="Process realtime", command=lambda: open_camera(selected_camera_index, selected_camera_index))
realtime_button.pack(side="left", padx=10, pady=10)

save_button = Button(app, text="Save image", command=lambda: save_current_frame(), state="disabled")
save_button.pack(side="left", padx=10, pady=10)

# start_button = Button(app, text="Start recording", command=lambda: start_recording(selected_camera_index))
# start_button.pack(side="left", padx=10, pady=10)

# stop_button = Button(app, text="Stop recording", command=lambda: stop_recording(selected_camera_index), state="disabled")
# stop_button.pack(side="left", padx=10, pady=10)

camera_menu = OptionMenu(app, StringVar(app, camera_options[1]), *camera_options, command=lambda index: select_camera(index))
camera_menu.pack(side="right", padx=10, pady=10)

brighten_menu = OptionMenu(app, StringVar(app, brighten_options[1]), *brighten_options, command=lambda index: select_brightness(index))
brighten_menu.pack(side="right", padx=10, pady=10)

image_button_basic = Button(app, text="Process images", command=lambda: open_image(Models.NAIVE, selected_camera_index, selected_brightness_index))
image_button_basic.pack(side="right", padx=10, pady=10)

video_button_basic = Button(app, text="Process video", command=lambda: open_video(Models.NAIVE, selected_camera_index, selected_brightness_index))
video_button_basic.pack(side="right", padx=10, pady=10)

# image_button_neural_network = Button(app, text="Process an image (neural network)", command=lambda: open_image(Models.SAM_FINETUNE, selected_camera_index, selected_brightness_index))
# image_button_neural_network.pack(side="right", padx=10, pady=10)

toggle_button = Button(app, text="Toggle processing", command=toggle_processing)
toggle_button.pack(side="right", padx=10, pady=10)

def only_numbers_validation_callback(P):
    if str.isdigit(P) or P == "": return True
    return False
    
numbers_validation = (app.register(only_numbers_validation_callback))

exposure_entry = Entry(app, width=10, validate='all', validatecommand=(numbers_validation, '%P'))
exposure_entry.pack(side="right", padx=5)
exposure_entry.insert(0, 5000)
exposure_label = Label(app, text="Exposure time (59 - 1000000 µs):")
exposure_label.pack(side="right", padx=5)

DEFAULT_FPS_VALUE = 60

fps_entry = Entry(app, width=10, validate='all', validatecommand=(numbers_validation, '%P'))
fps_entry.pack(side="right", padx=5)
fps_entry.insert(0, DEFAULT_FPS_VALUE)
fps_label = Label(app, text=f"FPS: {DEFAULT_FPS_VALUE}")
fps_label.pack(side="right", padx=5)

scale_entry = Entry(app, width=10, validate='all', validatecommand=(numbers_validation, '%P'))
scale_entry.pack(side="right", padx=5)
scale_entry.insert(0, DEFAULT_CONVERSION_SCALE)
scale_label = Label(app, text="800 pixels to µm:")
scale_label.pack(side="right", padx=5)

app.mainloop()
